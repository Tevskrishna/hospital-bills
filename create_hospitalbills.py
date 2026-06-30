"""
Mallareddy Hospital bills — Venkateswara Rao (2026)
Excel workbook with dashboard charts + WhatsApp daily report sheet.
"""
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, BarChart3D, LineChart, PieChart, PieChart3D, Reference,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, datetime

OUTPUT = r"c:\Users\Admin\Desktop\Hospital bills\hospitalbills.xlsx"

BILLS_DATA_START = 5
BILLS_EXTRA_ROWS = 60
ADV_DATA_START = 4
ADV_EXTRA_ROWS = 25

# (date, paid_by, amount, payment_mode, notes)
HOSPITAL_BILLS = [
    (date(2026, 6, 28), "Venky", 50, "", ""),
    (date(2026, 6, 28), "Venky", 733, "", ""),
    (date(2026, 6, 28), "Venky", 300, "", ""),
    (date(2026, 6, 29), "Venky", 500, "", ""),
    (date(2026, 6, 29), "Venky", 68.16, "", ""),
    (date(2026, 6, 29), "Venky", 250, "", ""),
    (date(2026, 6, 29), "Venky", 888, "", ""),
    (date(2026, 6, 29), "Venky", 1300, "", ""),
    (date(2026, 6, 30), "Venky", 95, "", ""),
    (date(2026, 6, 30), "Venky", 888, "", ""),
    (date(2026, 6, 30), "Venky", 908.28, "", ""),
    (date(2026, 6, 28), "Deepa", 1100, "", "Rajini's wife"),
    (date(2026, 6, 28), "Deepa", 110.66, "", ""),
    (date(2026, 6, 28), "Deepa", 550, "", ""),
    (date(2026, 6, 28), "Deepa", 22.86, "", ""),
    (date(2026, 6, 28), "Deepa", 500, "", ""),
    (date(2026, 6, 28), "Deepa", 58, "", ""),
    (date(2026, 6, 28), "Deepa", 800, "", ""),
    (date(2026, 6, 28), "Deepa", 1500, "", ""),
    (date(2026, 6, 28), "Deepa", 800, "", ""),
    (date(2026, 6, 29), "Deepa", 1051.85, "", ""),
    (date(2026, 6, 29), "Deepa", 990, "", ""),
    (date(2026, 6, 29), "Deepa", 500, "", ""),
    (date(2026, 6, 29), "Deepa", 2500, "", ""),
    (date(2026, 6, 29), "Deepa", 60, "", ""),
    (date(2026, 6, 30), "Deepa", 800, "", ""),
    (date(2026, 6, 30), "Deepa", 90.66, "", ""),
    (date(2026, 6, 30), "Deepa", 1334.54, "", ""),
    (date(2026, 6, 30), "Deepa", 1850, "Cash", "Paid in cash"),
    (date(2026, 6, 30), "Kalyan", 3600, "Credit Card", "Single payment via credit card"),
]

SHIVAJI_TO_VENKY = [
    (date(2026, 6, 28), 3009, "Advance to Venky for bills"),
    (date(2026, 6, 29), 2000, "Advance to Venky for bills"),
    (date(2026, 6, 30), 1000, "Advance to Venky for bills"),
    (date(2026, 6, 30), 1000, "Advance to Venky for bills"),
    (date(2026, 6, 30), 500, "Advance to Venky for bills"),
]

# Power BI–inspired palette
C_NAVY = "0D1B2A"
C_BLUE = "1B4965"
C_TEAL = "2A9D8F"
C_GOLD = "E9C46A"
C_CORAL = "E76F51"
C_LIGHT = "F8F9FA"
C_CARD1 = "1B4965"
C_CARD2 = "2A9D8F"
C_CARD3 = "E76F51"
C_CARD4 = "264653"
CHART_COLORS = ["2A9D8F", "E76F51", "E9C46A", "1B4965", "F4A261"]

MONEY_FMT = '#,##0.00'
DATE_FMT = "DD-MMM-YYYY"
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def bills_end():
    return BILLS_DATA_START + len(HOSPITAL_BILLS) + BILLS_EXTRA_ROWS - 1


def adv_end():
    return ADV_DATA_START + len(SHIVAJI_TO_VENKY) + ADV_EXTRA_ROWS - 1


def bills_range(col):
    letter = get_column_letter(col)
    return f"'Hospital Bills'!${letter}${BILLS_DATA_START}:${letter}${bills_end()}"


def adv_range(col):
    letter = get_column_letter(col)
    return f"'Shivaji Advances'!${letter}${ADV_DATA_START}:${letter}${adv_end()}"


def hdr(ws, row, cols, fill=C_NAVY):
    f = PatternFill("solid", fgColor=fill)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def cell_border(cell, alt=False):
    cell.border = BORDER
    if alt:
        cell.fill = PatternFill("solid", fgColor="EDF2F7")


def autosize(ws, mn=10, mx=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[letter].width = min(max(w, mn), mx)


def style_chart(chart, title, w=18, h=11):
    chart.title = title
    chart.style = 10
    chart.width = w
    chart.height = h
    chart.legend.position = "b"
    chart.legend.overlay = False
    if hasattr(chart, "y_axis") and chart.y_axis:
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = None
    if hasattr(chart, "x_axis") and chart.x_axis:
        chart.x_axis.delete = False


def color_bar_series(chart, colors):
    if not chart.series:
        return
    series = chart.series[0]
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)


def build_bills_sheet(wb):
    ws = wb.active
    ws.title = "Hospital Bills"
    ws["A1"] = "Mallareddy Hospital — Sri Venkateswara Rao"
    ws["A1"].font = Font(bold=True, size=15, color=C_NAVY)
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Add new payments in empty rows. "
        "Cumulative Spend = running total of ALL hospital bills from row 1 till this row."
    )
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:G2")

    headers = [
        "S.No", "Date", "Paid By", "Amount (₹)", "Payment Mode",
        "Notes", "Cumulative Spend (₹)",
    ]
    r0 = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=r0, column=i, value=h)
    hdr(ws, r0, len(headers))

    for idx, row_data in enumerate(HOSPITAL_BILLS, 1):
        r = BILLS_DATA_START + idx - 1
        dt, payer, amt, mode, notes = row_data
        alt = idx % 2 == 0
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=dt).number_format = DATE_FMT
        ws.cell(row=r, column=3, value=payer)
        ws.cell(row=r, column=4, value=amt).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=mode)
        ws.cell(row=r, column=6, value=notes)
        for c in range(1, 7):
            cell_border(ws.cell(row=r, column=c), alt)

    for i in range(len(HOSPITAL_BILLS), len(HOSPITAL_BILLS) + BILLS_EXTRA_ROWS):
        r = BILLS_DATA_START + i
        ws.cell(row=r, column=1, value=i + 1)
        for c in range(1, 7):
            cell_border(ws.cell(row=r, column=c), i % 2 == 0)
        ws.cell(row=r, column=4).number_format = MONEY_FMT

    end = bills_end()
    for r in range(BILLS_DATA_START, end + 1):
        if r == BILLS_DATA_START:
            f = f'=IF(D{r}="","",D{r})'
        else:
            f = f'=IF(D{r}="","",G{r-1}+D{r})'
        ws.cell(row=r, column=7, value=f).number_format = MONEY_FMT
        cell_border(ws.cell(row=r, column=7), (r - BILLS_DATA_START) % 2 == 1)

    ref = f"A{r0}:G{end}"
    tab = Table(displayName="HospitalBills", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)

    sr = end + 3
    ws.cell(row=sr, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12, color=C_NAVY)
    ws.cell(row=sr, column=4, value=f"=SUM(D{BILLS_DATA_START}:D{end})")
    ws.cell(row=sr, column=4).number_format = MONEY_FMT
    ws.cell(row=sr, column=4).font = Font(bold=True, size=12, color=C_CORAL)

    for i, p in enumerate(["Venky", "Deepa", "Kalyan"]):
        r = sr + 2 + i
        ws.cell(row=r, column=1, value=f"Total — {p}")
        ws.cell(row=r, column=4, value=f'=SUMIF(C{BILLS_DATA_START}:C{end},"{p}",D{BILLS_DATA_START}:D{end})')
        ws.cell(row=r, column=4).number_format = MONEY_FMT

    ws.conditional_formatting.add(
        f"D{BILLS_DATA_START}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=C_TEAL, showValue=True, minLength=0, maxLength=100),
    )
    ws.freeze_panes = f"A{BILLS_DATA_START}"
    autosize(ws)


def build_advances_sheet(wb):
    ws = wb.create_sheet("Shivaji Advances")
    ws["A1"] = "Shivaji → Venky (funds for hospital bill management)"
    ws["A1"].font = Font(bold=True, size=14, color=C_NAVY)
    ws.merge_cells("A1:E1")

    headers = ["S.No", "Date", "From", "Amount (₹)", "Notes"]
    r0 = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=r0, column=i, value=h)
    hdr(ws, r0, len(headers))

    for idx, (dt, amt, notes) in enumerate(SHIVAJI_TO_VENKY, 1):
        r = ADV_DATA_START + idx - 1
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=dt).number_format = DATE_FMT
        ws.cell(row=r, column=3, value="Shivaji → Venky")
        ws.cell(row=r, column=4, value=amt).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=notes)
        for c in range(1, 6):
            cell_border(ws.cell(row=r, column=c), idx % 2 == 0)

    for i in range(len(SHIVAJI_TO_VENKY), len(SHIVAJI_TO_VENKY) + ADV_EXTRA_ROWS):
        r = ADV_DATA_START + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=3, value="Shivaji → Venky")
        for c in range(1, 6):
            cell_border(ws.cell(row=r, column=c))

    end = adv_end()
    ref = f"A{r0}:E{end}"
    tab = Table(displayName="ShivajiAdvances", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)

    ws.cell(row=end + 2, column=1, value="Total from Shivaji").font = Font(bold=True)
    ws.cell(row=end + 2, column=4, value=f"=SUM(D{ADV_DATA_START}:D{end})")
    ws.cell(row=end + 2, column=4).number_format = MONEY_FMT
    ws.cell(row=end + 2, column=4).font = Font(bold=True, size=12)
    ws.freeze_panes = f"A{ADV_DATA_START}"
    autosize(ws)


def build_chart_data(wb):
    """Hidden helper sheet — feeds dashboard charts via live formulas."""
    ws = wb.create_sheet("_ChartData")
    ws.sheet_state = "hidden"
    be, ae = bills_end(), adv_end()

    ws["A1"] = "Person"
    ws["B1"] = "Amount"
    persons = ["Venky", "Deepa", "Kalyan"]
    for i, p in enumerate(persons, 2):
        ws.cell(row=i, column=1, value=p)
        ws.cell(row=i, column=2, value=f"=SUMIF({bills_range(3)},\"{p}\",{bills_range(4)})")
        ws.cell(row=i, column=2).number_format = MONEY_FMT

    ws["D1"] = "Date"
    ws["E1"] = "Venky"
    ws["F1"] = "Deepa"
    ws["G1"] = "Kalyan"
    ws["H1"] = "Day Total"
    dates = sorted({d for d, *_ in HOSPITAL_BILLS})
    for i, dt in enumerate(dates, 2):
        ws.cell(row=i, column=4, value=dt).number_format = DATE_FMT
        for j, person in enumerate(persons, 5):
            col_letter = get_column_letter(j)
            ws.cell(
                row=i, column=j,
                value=(
                    f"=SUMIFS({bills_range(4)},{bills_range(2)},D{i},"
                    f"{bills_range(3)},\"{person}\")"
                ),
            ).number_format = MONEY_FMT
        ws.cell(row=i, column=8, value=f"=SUM(E{i}:G{i})").number_format = MONEY_FMT

    last_date_row = 1 + len(dates)
    ws["J1"] = "Payment Mode"
    ws["K1"] = "Amount"
    modes = ["Cash", "Credit Card", ""]
    labels = ["Cash", "Credit Card", "Other / Not specified"]
    for i, (mode, label) in enumerate(zip(modes, labels), 2):
        ws.cell(row=i, column=10, value=label)
        if mode:
            ws.cell(
                row=i, column=11,
                value=f"=SUMIF({bills_range(5)},\"{mode}\",{bills_range(4)})",
            )
        else:
            ws.cell(
                row=i, column=11,
                value=f"=SUM({bills_range(4)})-SUMIF({bills_range(5)},\"Cash\",{bills_range(4)})-SUMIF({bills_range(5)},\"Credit Card\",{bills_range(4)})",
            )
        ws.cell(row=i, column=11).number_format = MONEY_FMT

    ws["M1"] = "Label"
    ws["N1"] = "Value"
    kpis = [
        ("Total Spend", f"=SUM({bills_range(4)})"),
        ("Shivaji → Venky", f"=SUM({adv_range(4)})"),
        ("Venky paid", f"=SUMIF({bills_range(3)},\"Venky\",{bills_range(4)})"),
        ("Deepa paid", f"=SUMIF({bills_range(3)},\"Deepa\",{bills_range(4)})"),
        ("Kalyan paid", f"=SUMIF({bills_range(3)},\"Kalyan\",{bills_range(4)})"),
        ("Venky balance*", f"=SUM({adv_range(4)})-SUMIF({bills_range(3)},\"Venky\",{bills_range(4)})"),
    ]
    for i, (lbl, formula) in enumerate(kpis, 2):
        ws.cell(row=i, column=13, value=lbl)
        ws.cell(row=i, column=14, value=formula).number_format = MONEY_FMT

    return last_date_row


def kpi_card(ws, row, col, label, formula, bg, wide=3):
    fill = PatternFill("solid", fgColor=bg)
    r, c = row, col
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + wide - 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + wide - 1)
    t = ws.cell(row=r, column=c, value=label)
    t.fill = fill
    t.font = Font(bold=True, color="FFFFFF", size=9)
    t.alignment = Alignment(horizontal="center", vertical="center")
    v = ws.cell(row=r + 1, column=c, value=formula)
    v.fill = PatternFill("solid", fgColor="FFFFFF")
    v.font = Font(bold=True, size=16, color=C_NAVY)
    v.number_format = MONEY_FMT
    v.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (r, r + 1):
        for cc in range(c, c + wide):
            ws.cell(row=rr, column=cc).border = Border(
                left=Side(style="medium", color=bg),
                right=Side(style="medium", color=bg),
                top=Side(style="medium", color=bg),
                bottom=Side(style="medium", color=bg),
            )


def build_dashboard(wb, last_date_row):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_TEAL

    for letter, w in [("A", 4), ("B", 14), ("C", 14), ("D", 14), ("E", 14), ("F", 4)]:
        ws.column_dimensions[letter].width = w

    ws.merge_cells("B2:E2")
    ws["B2"] = "HOSPITAL BILLS DASHBOARD"
    ws["B2"].font = Font(bold=True, size=22, color=C_NAVY)
    ws.merge_cells("B3:E3")
    ws["B3"] = "Sri Venkateswara Rao  •  Mallareddy Hospital  •  Year 2026"
    ws["B3"].font = Font(size=11, color="64748B")
    ws["F2"] = "=TODAY()"
    ws["F2"].number_format = "DD-MMM-YYYY"
    ws["F2"].font = Font(bold=True, color=C_TEAL)
    ws["F2"].alignment = Alignment(horizontal="right")

    be = bills_end()
    ae = adv_end()
    kpi_card(ws, 5, 2, "TOTAL HOSPITAL SPEND", f"=SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be})", C_CARD1, wide=1)
    kpi_card(ws, 5, 3, "VENKY PAID", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})", C_CARD2, wide=1)
    kpi_card(ws, 5, 4, "DEEPA PAID (Rajini)", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Deepa\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})", C_CARD3, wide=1)
    kpi_card(ws, 5, 5, "KALYAN PAID", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Kalyan\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})", C_CARD4, wide=1)

    kpi_card(
        ws, 8, 2, "SHIVAJI → VENKY (advances)",
        f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})", "5C4D7D", wide=2,
    )
    kpi_card(
        ws, 8, 4, "VENKY BALANCE vs SHIVAJI*",
        f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})-SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})",
        C_TEAL, wide=2,
    )

    ws.merge_cells("B11:E11")
    ws["B11"] = (
        "* Venky Balance = money Shivaji gave Venky minus what Venky paid at hospital. "
        "Positive = Shivaji still owes Venky; Negative = Venky has extra from Shivaji."
    )
    ws["B11"].font = Font(italic=True, size=8, color="94A3B8")

    # --- Charts from _ChartData ---
    cd = wb["_ChartData"]

    # 3D Bar — spend by person
    bar3d = BarChart3D()
    bar3d.type = "col"
    style_chart(bar3d, "Spend by Person (3D)", w=16, h=12)
    data = Reference(cd, min_col=2, min_row=1, max_row=4)
    cats = Reference(cd, min_col=1, min_row=2, max_row=4)
    bar3d.add_data(data, titles_from_data=True)
    bar3d.set_categories(cats)
    color_bar_series(bar3d, CHART_COLORS[:3])
    ws.add_chart(bar3d, "B13")

    # Pie — share %
    pie = PieChart()
    style_chart(pie, "Share of Total Spend", w=14, h=12)
    pie_data = Reference(cd, min_col=2, min_row=2, max_row=4)
    pie_cats = Reference(cd, min_col=1, min_row=2, max_row=4)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    color_bar_series(pie, CHART_COLORS[:3])
    ws.add_chart(pie, "H13")

    # Stacked column — date wise
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    style_chart(bar, "Daily Spend by Person", w=20, h=12)
    ddata = Reference(cd, min_col=5, min_row=1, max_col=7, max_row=last_date_row)
    dcats = Reference(cd, min_col=4, min_row=2, max_row=last_date_row)
    bar.add_data(ddata, titles_from_data=True)
    bar.set_categories(dcats)
    ws.add_chart(bar, "B30")

    # 3D Pie — payment mode
    pie3d = PieChart3D()
    style_chart(pie3d, "Payment Mode Split", w=14, h=12)
    pm_data = Reference(cd, min_col=11, min_row=2, max_row=4)
    pm_cats = Reference(cd, min_col=10, min_row=2, max_row=4)
    pie3d.add_data(pm_data, titles_from_data=False)
    pie3d.set_categories(pm_cats)
    pie3d.dataLabels = DataLabelList()
    pie3d.dataLabels.showPercent = True
    color_bar_series(pie3d, ["E76F51", "2A9D8F", "E9C46A"])
    ws.add_chart(pie3d, "H30")

    # Line — cumulative spend by date
    line = LineChart()
    style_chart(line, "Cumulative Hospital Spend Trend", w=20, h=12)
    cd["I1"] = "Cumulative"
    dates_sorted = sorted({d for d, *_ in HOSPITAL_BILLS})
    for i in range(2, 2 + len(dates_sorted)):
        cd.cell(row=i, column=9, value=f"=SUM($H$2:H{i})").number_format = MONEY_FMT
    line_data = Reference(cd, min_col=9, min_row=1, max_row=1 + len(dates_sorted))
    line_cats = Reference(cd, min_col=4, min_row=2, max_row=1 + len(dates_sorted))
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_cats)
    ws.add_chart(line, "B47")

    ws.merge_cells("B62:U62")
    ws["B62"] = "Screenshot this sheet for family WhatsApp group — or use the 'WhatsApp Report' tab."
    ws["B62"].font = Font(italic=True, size=9, color=C_TEAL)


def build_date_summary(wb):
    ws = wb.create_sheet("Date-wise Summary")
    ws["A1"] = "Spend by date"
    ws["A1"].font = Font(bold=True, size=14, color=C_NAVY)
    headers = ["Date", "Venky (₹)", "Deepa (₹)", "Kalyan (₹)", "Day Total (₹)", "% of Grand Total"]
    r0 = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=r0, column=i, value=h)
    hdr(ws, r0, len(headers))

    dates = sorted({d for d, *_ in HOSPITAL_BILLS})
    be = bills_end()
    for idx, dt in enumerate(dates):
        r = r0 + 1 + idx
        ws.cell(row=r, column=1, value=dt).number_format = DATE_FMT
        for col, person in enumerate(["Venky", "Deepa", "Kalyan"], 2):
            ws.cell(
                row=r, column=col,
                value=(
                    f"=SUMIFS('Hospital Bills'!$D${BILLS_DATA_START}:$D${be},"
                    f"'Hospital Bills'!$B${BILLS_DATA_START}:$B${be},A{r},"
                    f"'Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"{person}\")"
                ),
            ).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = MONEY_FMT
        ws.cell(
            row=r, column=6,
            value=f"=IF(SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be})=0,\"\",E{r}/SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be}))",
        ).number_format = "0.0%"
        for c in range(1, 7):
            cell_border(ws.cell(row=r, column=c), idx % 2 == 0)

    tr = r0 + 1 + len(dates)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 6):
        L = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f"=SUM({L}{r0+1}:{L}{tr-1})")
        ws.cell(row=tr, column=c).number_format = MONEY_FMT if c < 6 else MONEY_FMT
    ws.cell(row=tr, column=6, value="100%").font = Font(bold=True)
    autosize(ws)


def build_settlement_summary(wb):
    ws = wb.create_sheet("Settlement")
    ws["A1"] = "Who paid what — settlement view"
    ws["A1"].font = Font(bold=True, size=14, color=C_NAVY)
    ws["A2"] = "Deepa = Rajini's wife  |  Venky manages Shivaji's share via advances"
    ws.merge_cells("A2:E2")
    ws["A2"].font = Font(size=9, color="64748B")

    be, ae = bills_end(), adv_end()
    rows = [
        ("Venky", "Paid at hospital", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Deepa (Rajini's wife)", "Paid at hospital", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Deepa\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Kalyan", "Paid at hospital", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Kalyan\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Shivaji", "Gave Venky (not hospital direct)", f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})"),
        ("Venky", "Balance vs Shivaji (+ owes Venky)", f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})-SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
    ]
    r0 = 4
    for i, h in enumerate(["Person", "Type", "Amount (₹)", "Share %"], 1):
        ws.cell(row=r0, column=i, value=h)
    hdr(ws, r0, 4)

    for i, (person, typ, formula) in enumerate(rows):
        r = r0 + 1 + i
        ws.cell(row=r, column=1, value=person)
        ws.cell(row=r, column=2, value=typ)
        ws.cell(row=r, column=3, value=formula).number_format = MONEY_FMT
        if i < 3:
            ws.cell(
                row=r, column=4,
                value=f"=IF(SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be})=0,\"\",C{r}/SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be}))",
            ).number_format = "0.0%"
        for c in range(1, 5):
            cell_border(ws.cell(row=r, column=c), i % 2 == 0)

    gr = r0 + 1 + len(rows) + 1
    ws.cell(row=gr, column=1, value="GRAND TOTAL (hospital only)").font = Font(bold=True, size=12)
    ws.cell(row=gr, column=3, value=f"=SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be})")
    ws.cell(row=gr, column=3).number_format = MONEY_FMT
    ws.cell(row=gr, column=3).font = Font(bold=True, color=C_CORAL, size=12)
    autosize(ws, mx=36)


def build_whatsapp_report(wb):
    ws = wb.create_sheet("WhatsApp Report")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "25D366"

    be, ae = bills_end(), adv_end()
    title_fill = PatternFill("solid", fgColor=C_NAVY)
    section_fill = PatternFill("solid", fgColor=C_TEAL)

    def banner(row, text, merge_to=6, fill=title_fill):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28

    banner(1, "🏥 MALLAREDDY HOSPITAL — DAILY BILL UPDATE")
    ws.merge_cells("A2:F2")
    ws["A2"] = "Patient: Sri Venkateswara Rao  |  Family: Shivaji • Rajini • Kalyan"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(size=10, color="475569")
    ws["A3"] = 'Report date:'
    ws["A3"].font = Font(bold=True, size=10)
    ws["B3"] = "=TODAY()"
    ws["B3"].number_format = "DD-MMM-YYYY"
    ws.merge_cells("B3:F3")
    ws["B3"].alignment = Alignment(horizontal="left")

    banner(5, "SUMMARY (auto-calculated)", fill=section_fill)
    summary = [
        ("Total hospital spend till date", f"=SUM('Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Venky paid", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Deepa paid (Rajini's wife)", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Deepa\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Kalyan paid (incl. ₹3600 credit card)", f"=SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Kalyan\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
        ("Shivaji → Venky (advances)", f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})"),
        ("Venky balance vs Shivaji*", f"=SUM('Shivaji Advances'!$D${ADV_DATA_START}:$D${ae})-SUMIF('Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"Venky\",'Hospital Bills'!$D${BILLS_DATA_START}:$D${be})"),
    ]
    for i, (lbl, fml) in enumerate(summary, 7):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        ws.cell(row=i, column=1, value=lbl).font = Font(size=10)
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
        v = ws.cell(row=i, column=4, value=fml)
        v.number_format = '₹ #,##0.00'
        v.font = Font(bold=True, size=11, color=C_NAVY)
        v.alignment = Alignment(horizontal="right")
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = BORDER

    banner(14, "DATE-WISE BREAKDOWN", fill=section_fill)
    hdr(ws, 15, 5)
    ws.cell(row=15, column=1, value="Date")
    ws.cell(row=15, column=2, value="Venky")
    ws.cell(row=15, column=3, value="Deepa")
    ws.cell(row=15, column=4, value="Kalyan")
    ws.cell(row=15, column=5, value="Day Total")

    dates = sorted({d for d, *_ in HOSPITAL_BILLS})
    for idx, dt in enumerate(dates):
        r = 16 + idx
        ws.cell(row=r, column=1, value=dt).number_format = DATE_FMT
        for col, person in enumerate(["Venky", "Deepa", "Kalyan"], 2):
            ws.cell(
                row=r, column=col,
                value=(
                    f"=SUMIFS('Hospital Bills'!$D${BILLS_DATA_START}:$D${be},"
                    f"'Hospital Bills'!$B${BILLS_DATA_START}:$B${be},A{r},"
                    f"'Hospital Bills'!$C${BILLS_DATA_START}:$C${be},\"{person}\")"
                ),
            ).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = MONEY_FMT

    vr = 16 + len(dates) + 2
    banner(vr, "PLEASE VERIFY ✓  Reply OK or send corrections", fill=PatternFill("solid", fgColor=C_CORAL))
    checks = [
        "☐ Venky — confirm your payments & Shivaji advances",
        "☐ Deepa — confirm amounts (₹1,850 was CASH on 30-Jun)",
        "☐ Kalyan — confirm ₹3,600 CREDIT CARD on 30-Jun",
        "☐ Shivaji — confirm advances to Venky: ₹7,509 total",
    ]
    for i, txt in enumerate(checks, vr + 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.cell(row=i, column=1, value=txt).font = Font(size=10)

    ws.merge_cells(start_row=vr + 6, start_column=1, end_row=vr + 6, end_column=6)
    ws.cell(
        row=vr + 6, column=1,
        value="* Cumulative Spend on Bills sheet = total of ALL bills added line-by-line (not per person).",
    ).font = Font(italic=True, size=8, color="94A3B8")

    for letter, w in [("A", 14), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 8)]:
        ws.column_dimensions[letter].width = w

    ws.print_area = f"A1:F{vr + 6}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


def main():
    wb = Workbook()
    build_bills_sheet(wb)
    build_advances_sheet(wb)
    last_dr = build_chart_data(wb)
    build_dashboard(wb, last_dr)
    build_date_summary(wb)
    build_settlement_summary(wb)
    build_whatsapp_report(wb)

    order = [
        "Dashboard", "WhatsApp Report", "Hospital Bills", "Date-wise Summary",
        "Settlement", "Shivaji Advances", "_ChartData",
    ]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
